"""Justificatif d'un remboursement : PDF et tableur.

C'est la piece que la comptabilite archive et qui justifie le virement. Elle
reprend la disposition du modele fourni par le client (« NDF - Nom Prenom ») :
identite du benevole, une ligne par note de frais, les totaux, puis le bloc
« Apurement » qui dit comment le versement a ete fait.

**Le tableur est reconstruit, pas rempli.** `openpyxl` perd graphiques, liens
externes et parametres d'impression a la reecriture d'un classeur — et le modele
en contient. Un fichier propre reprenant libelles et disposition vaut mieux
qu'un modele silencieusement mutile.

Module volontairement pur : il ne connait ni la base ni FastAPI, ce qui le rend
testable sur des objets quelconques et reutilisable ailleurs (un export mensuel,
par exemple).
"""

from __future__ import annotations

import io
from dataclasses import dataclass
from datetime import date
from decimal import Decimal
from typing import Any, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    KeepTogether,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from app.core.errors import ErrorCode
from app.core.exceptions import AppException
from app.core.money import montant_du


# Libelles repris du modele, au mot pres : la comptabilite les reconnait d'un
# coup d'oeil, et les rapprocher d'un fichier a l'autre doit rester immediat.
ENTETES = (
    "Date",
    "Rattachement (évènement, activité..)",
    "Montant",
    "Fournisseur",
    "Nature charge",
    "Commentaires (Nombre de repas, …)",
)
MENTION_MODELE = (
    "NB : La note de frais est à joindre avec ce fichier en format EXCEL "
    "(détaillé de toutes les factures avec son rattachement) ainsi que "
    "l'ensemble des justificatifs."
)
TITRE_APUREMENT = "Apurement (partie réservée KOUTTAB) :"

# Premiere ligne de donnees dans le tableur : le modele place ses en-tetes en
# ligne 10, apres le bloc d'identite.
LIGNE_ENTETES = 10
LIGNE_PREMIERE_NOTE = LIGNE_ENTETES + 1

_VERT = colors.HexColor("#3d4f3d")  # charte de l'institut
_SABLE = colors.HexColor("#d7bd9a")


@dataclass(frozen=True)
class DonneesJustificatif:
    """Tout ce qu'il faut pour produire le document, sans toucher a la base."""

    benevole: Any
    notes: Sequence[Any]
    date_remboursement: date
    moyen: str
    etablissement: str
    approuve_par: str
    commentaire: str | None = None


# ---------------------------------------------------------------------------
# Mise en forme partagee
# ---------------------------------------------------------------------------


def _euros(valeur: Decimal | None) -> str:
    montant = Decimal(valeur or 0)
    return f"{montant:,.2f} €".replace(",", " ").replace(".", ",")


def _date_fr(valeur: date | None) -> str:
    return valeur.strftime("%d/%m/%Y") if valeur else ""


def _rattachement(note: Any) -> str:
    """Meme arbitrage qu'au depot : evenement, sinon categorie.

    `rattachement` ferme la marche pour les notes anterieures au module
    comptable, qui ne portent que ce champ libre.
    """
    return (
        getattr(note, "evenement", None)
        or getattr(note, "categorie", None)
        or getattr(note, "rattachement", None)
        or ""
    )


def _identite(benevole: Any) -> tuple[str, str, str, str]:
    return (
        getattr(benevole, "nom", None) or "",
        getattr(benevole, "prenom", None) or "",
        getattr(benevole, "email", None) or "",
        getattr(benevole, "telephone", None) or "",
    )


def _totaux(notes: Sequence[Any]) -> dict[str, Decimal]:
    """Sous-total, avances, remises et net a payer.

    Le sous-total est la somme des montants **bruts** : le document doit montrer
    ce qui a ete depense avant d'expliquer ce qui est verse. Le TOTAL, lui,
    passe par `montant_du`, qui borne chaque note a zero.
    """
    brut = sum((Decimal(n.montant or 0) for n in notes), Decimal("0"))
    avances = sum(
        (Decimal(getattr(n, "remboursement_deja_emis", 0) or 0) for n in notes),
        Decimal("0"),
    )
    remises = sum((Decimal(getattr(n, "remise", 0) or 0) for n in notes), Decimal("0"))
    net = sum((montant_du(n) for n in notes), Decimal("0"))
    return {"brut": brut, "avances": avances, "remises": remises, "net": net}


def _verifier(donnees: DonneesJustificatif) -> None:
    if not donnees.notes:
        raise AppException(
            ErrorCode.VALIDATION_ERROR,
            detail="Aucune note de frais : il n'y a rien a justifier.",
        )


def nom_document(donnees: DonneesJustificatif) -> str:
    """Nom lisible du document, dans le style du modele du client.

    Sert de titre DANS le fichier. Le nom du fichier envoye par courriel passe,
    lui, par `services/naming.py` : il doit rester en ASCII pur, sans quoi les
    clients de messagerie deforment le nom de la piece jointe.
    """
    nom, prenom, _, _ = _identite(donnees.benevole)
    identite = " ".join(p for p in (nom, prenom) if p) or getattr(
        donnees.benevole, "username", "Benevole"
    )
    return f"NDF - {identite}"


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------


def construire_pdf(donnees: DonneesJustificatif) -> bytes:
    """Rend le justificatif en PDF A4."""
    _verifier(donnees)

    tampon = io.BytesIO()
    document = SimpleDocTemplate(
        tampon,
        pagesize=A4,
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
        title=nom_document(donnees),
        author="Le Kouttâb",
    )

    styles = getSampleStyleSheet()
    style_titre = ParagraphStyle(
        "TitreNdf",
        parent=styles["Title"],
        fontSize=16,
        textColor=_VERT,
        alignment=0,
        spaceAfter=2 * mm,
    )
    style_normal = ParagraphStyle("Corps", parent=styles["Normal"], fontSize=9)
    style_petit = ParagraphStyle(
        "Petit", parent=styles["Normal"], fontSize=7.5, textColor=colors.grey
    )
    style_cellule = ParagraphStyle("Cellule", parent=styles["Normal"], fontSize=8)

    nom, prenom, mail, telephone = _identite(donnees.benevole)
    totaux = _totaux(donnees.notes)

    contenu: list[Any] = [
        Paragraph(nom_document(donnees), style_titre),
        Spacer(1, 3 * mm),
    ]

    identite = Table(
        [
            ["Nom", nom, "Date de remise NDF", _date_fr(donnees.date_remboursement)],
            ["Prénom", prenom, "", ""],
            ["Mail", mail, "", ""],
            ["Téléphone", telephone, "", ""],
        ],
        colWidths=[25 * mm, 65 * mm, 45 * mm, 45 * mm],
    )
    identite.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("FONTNAME", (2, 0), (2, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
            ]
        )
    )
    contenu += [identite, Spacer(1, 4 * mm), Paragraph(MENTION_MODELE, style_petit)]
    contenu.append(Spacer(1, 4 * mm))

    lignes: list[list[Any]] = [[Paragraph(f"<b>{e}</b>", style_cellule) for e in ENTETES]]
    for note in donnees.notes:
        lignes.append(
            [
                Paragraph(_date_fr(getattr(note, "date_depense", None)), style_cellule),
                Paragraph(_rattachement(note), style_cellule),
                Paragraph(_euros(getattr(note, "montant", None)), style_cellule),
                Paragraph(getattr(note, "fournisseur", None) or "", style_cellule),
                Paragraph(getattr(note, "nature_charge", None) or "", style_cellule),
                Paragraph(getattr(note, "commentaires", None) or "", style_cellule),
            ]
        )

    tableau = Table(
        lignes,
        colWidths=[20 * mm, 40 * mm, 22 * mm, 30 * mm, 28 * mm, 40 * mm],
        repeatRows=1,
    )
    tableau.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), _VERT),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.4, _SABLE),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ALIGN", (2, 1), (2, -1), "RIGHT"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#faf7f2")]),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]
        )
    )
    contenu += [tableau, Spacer(1, 4 * mm)]

    recap = Table(
        [
            ["Sous total :", _euros(totaux["brut"])],
            ["Remboursement déjà émis (espèce) :", _euros(totaux["avances"])],
            ["Remise (à préciser) :", _euros(totaux["remises"])],
            ["TOTAL :", _euros(totaux["net"])],
        ],
        colWidths=[70 * mm, 35 * mm],
        hAlign="RIGHT",
    )
    recap.setStyle(
        TableStyle(
            [
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                ("LINEABOVE", (0, -1), (-1, -1), 0.8, _VERT),
                ("TOPPADDING", (0, 0), (-1, -1), 2),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
            ]
        )
    )
    contenu += [recap, Spacer(1, 6 * mm)]

    apurement = Table(
        [
            [Paragraph(f"<b>{TITRE_APUREMENT}</b>", style_normal), ""],
            ["Remboursement émis le :", _date_fr(donnees.date_remboursement)],
            ["Moyen :", donnees.moyen],
            ["Etablissement :", donnees.etablissement],
            ["Approuvé par :", donnees.approuve_par],
        ],
        colWidths=[60 * mm, 60 * mm],
    )
    apurement.setStyle(
        TableStyle(
            [
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("SPAN", (0, 0), (1, 0)),
                ("BOX", (0, 0), (-1, -1), 0.6, _VERT),
                ("INNERGRID", (0, 1), (-1, -1), 0.3, _SABLE),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#eee4d3")),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]
        )
    )
    # `KeepTogether` : le bloc d'apurement coupe en deux pages serait illisible,
    # et c'est lui qui porte la preuve du versement.
    contenu.append(KeepTogether(apurement))

    if donnees.commentaire:
        contenu += [Spacer(1, 4 * mm), Paragraph(donnees.commentaire, style_petit)]

    document.build(contenu)
    return tampon.getvalue()


# ---------------------------------------------------------------------------
# Tableur
# ---------------------------------------------------------------------------


def construire_xlsx(donnees: DonneesJustificatif) -> bytes:
    """Rend le justificatif en classeur, sur la disposition du modele."""
    _verifier(donnees)

    classeur = Workbook()
    feuille = classeur.active
    feuille.title = donnees.date_remboursement.strftime("%m %y")

    gras = Font(bold=True)
    entete_fond = PatternFill("solid", fgColor="3D4F3D")
    entete_police = Font(bold=True, color="FFFFFF")
    bordure = Border(*(Side(style="thin", color="D7BD9A"),) * 4)
    droite = Alignment(horizontal="right")

    nom, prenom, mail, telephone = _identite(donnees.benevole)
    for ligne, (libelle, valeur) in enumerate(
        (("Nom", nom), ("Prénom", prenom), ("Mail", mail), ("Téléphone", telephone)),
        start=1,
    ):
        feuille.cell(row=ligne, column=1, value=libelle).font = gras
        feuille.cell(row=ligne, column=2, value=valeur)

    feuille.cell(row=6, column=1, value="Date de remise NDF").font = gras
    cellule_date = feuille.cell(row=6, column=3, value=donnees.date_remboursement)
    cellule_date.number_format = "DD/MM/YYYY"

    feuille.cell(row=8, column=1, value=MENTION_MODELE).font = Font(size=8, italic=True)

    for colonne, entete in enumerate(ENTETES, start=1):
        cellule = feuille.cell(row=LIGNE_ENTETES, column=colonne, value=entete)
        cellule.font = entete_police
        cellule.fill = entete_fond
        cellule.border = bordure
        cellule.alignment = Alignment(wrap_text=True, vertical="center")

    ligne = LIGNE_PREMIERE_NOTE
    for note in donnees.notes:
        valeurs = (
            getattr(note, "date_depense", None),
            _rattachement(note),
            # `float` et non `str` : un montant en texte interdirait toute somme
            # dans le tableur, ce qui est precisement ce qu'on en attend.
            float(Decimal(getattr(note, "montant", 0) or 0)),
            getattr(note, "fournisseur", None) or "",
            getattr(note, "nature_charge", None) or "",
            getattr(note, "commentaires", None) or "",
        )
        for colonne, valeur in enumerate(valeurs, start=1):
            cellule = feuille.cell(row=ligne, column=colonne, value=valeur)
            cellule.border = bordure
            if colonne == 1:
                cellule.number_format = "DD/MM/YYYY"
            elif colonne == 3:
                cellule.number_format = '#,##0.00 "€"'
        ligne += 1

    totaux = _totaux(donnees.notes)
    ligne += 1
    for libelle, valeur, en_gras in (
        ("Sous total :", totaux["brut"], False),
        ("Remboursement déjà émis (espèce) :", totaux["avances"], False),
        ("Remise (à préciser) :", totaux["remises"], False),
        ("TOTAL :", totaux["net"], True),
    ):
        cellule_libelle = feuille.cell(row=ligne, column=2, value=libelle)
        cellule_valeur = feuille.cell(row=ligne, column=3, value=float(valeur))
        cellule_valeur.number_format = '#,##0.00 "€"'
        cellule_valeur.alignment = droite
        if en_gras:
            cellule_libelle.font = gras
            cellule_valeur.font = gras
        ligne += 1

    ligne += 1
    feuille.cell(row=ligne, column=1, value=TITRE_APUREMENT).font = gras
    for libelle, valeur in (
        ("Remboursement émis le :", donnees.date_remboursement),
        ("Moyen :", donnees.moyen),
        ("Etablissement :", donnees.etablissement),
        ("Approuvé par :", donnees.approuve_par),
    ):
        ligne += 1
        feuille.cell(row=ligne, column=2, value=libelle)
        cellule = feuille.cell(row=ligne, column=3, value=valeur)
        if isinstance(valeur, date):
            cellule.number_format = "DD/MM/YYYY"

    for colonne, largeur in enumerate((12, 34, 14, 22, 20, 40), start=1):
        feuille.column_dimensions[get_column_letter(colonne)].width = largeur

    tampon = io.BytesIO()
    classeur.save(tampon)
    return tampon.getvalue()


__all__ = [
    "DonneesJustificatif",
    "construire_pdf",
    "construire_xlsx",
    "nom_document",
]
