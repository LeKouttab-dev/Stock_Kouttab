"""Aligne les quantites de la buvette sur un inventaire tenu dans un classeur.

L'inventaire de l'espace vente vivait dans un fichier Excel avant cette
application. Ce script en reprend la colonne de stock final et l'applique aux
produits de la buvette.

    sudo docker compose cp "inventaire.xlsx" api:/tmp/inv.xlsx
    sudo docker compose exec api python scripts/importer_stock_buvette.py \\
        --fichier /tmp/inv.xlsx

**Rien n'est ecrit sans `--appliquer`.** Sans ce drapeau, le script affiche ce
qu'il changerait et s'arrete la : un inventaire mal lu fausserait tout le stock,
et une correspondance de noms se relit avant, pas apres.

**Seules les quantites bougent.** Ni les prix, ni les seuils d'alerte, ni les
libelles : les produits viennent de HelloAsso, qui en reste la source.

Un produit du classeur absent de la buvette est **ignore et signale** — c'est la
consigne : la boutique HelloAsso fait foi sur ce qui existe.
"""

from __future__ import annotations

import argparse
import sys
import unicodedata
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import openpyxl  # noqa: E402
from sqlalchemy import select  # noqa: E402
from sqlalchemy.orm import Session  # noqa: E402

from app.core.logger import get_logger  # noqa: E402
from app.db.models import BuvetteProduct  # noqa: E402
from app.db.session import SessionLocal  # noqa: E402


logger = get_logger("import.stock")

# En-tetes de la feuille d'inventaire, tels qu'ils figurent en ligne 5 :
#   Categorie | Designation | P.U achat | P.U vente | (SI) | Achat | Vendu |
#   (SF, dater du jour) | Prix de revient | CA theorique | Marge
COLONNE_DESIGNATION = 2
COLONNE_STOCK_FINAL = 8
PREMIERE_LIGNE = 6


def normaliser(libelle: str) -> str:
    """Ramene un libelle a une forme comparable.

    Les noms sont saisis a la main dans le classeur et par HelloAsso dans la
    boutique : accents, majuscules et espaces multiples divergent forcement.
    Comparer les chaines brutes ne rapprocherait presque rien.
    """
    sans_accents = "".join(
        c for c in unicodedata.normalize("NFD", libelle) if unicodedata.category(c) != "Mn"
    )
    return " ".join(sans_accents.lower().replace("'", "'").split())


def lire_inventaire(chemin: Path, feuille: str | None) -> list[tuple[str, int]]:
    """Rend les couples (designation, stock final) reellement renseignes.

    Une cellule de stock vide n'est pas un zero : elle signifie que le produit
    n'a pas ete compte. L'ecrire a zero effacerait un stock reel.
    """
    classeur = openpyxl.load_workbook(chemin, data_only=True)
    nom = feuille or classeur.sheetnames[-1]
    if nom not in classeur.sheetnames:
        raise SystemExit(
            f"Feuille '{nom}' introuvable. Feuilles disponibles : "
            + ", ".join(classeur.sheetnames)
        )
    onglet = classeur[nom]
    print(f"Feuille lue : {nom}\n")

    releve: list[tuple[str, int]] = []
    ignores_sans_stock: list[str] = []
    for ligne in onglet.iter_rows(
        min_row=PREMIERE_LIGNE, max_row=onglet.max_row, max_col=COLONNE_STOCK_FINAL,
        values_only=True,
    ):
        designation = ligne[COLONNE_DESIGNATION - 1]
        stock = ligne[COLONNE_STOCK_FINAL - 1]
        if not designation or not str(designation).strip():
            continue
        if stock is None or str(stock).strip() == "":
            ignores_sans_stock.append(str(designation).strip())
            continue
        try:
            quantite = int(round(float(stock)))
        except (TypeError, ValueError):
            ignores_sans_stock.append(f"{designation} (stock illisible : {stock!r})")
            continue
        releve.append((str(designation).strip(), max(quantite, 0)))

    if ignores_sans_stock:
        print(f"Sans stock compte dans le classeur — laisses tels quels ({len(ignores_sans_stock)}) :")
        for nom_produit in ignores_sans_stock:
            print(f"  · {nom_produit}")
        print()
    return releve


def appliquer(db: Session, releve: list[tuple[str, int]], *, ecrire: bool) -> int:
    produits = list(db.execute(select(BuvetteProduct)).scalars().all())
    par_nom = {normaliser(p.name): p for p in produits}

    changements = 0
    inchanges = 0
    absents: list[str] = []

    for designation, quantite in releve:
        produit = par_nom.get(normaliser(designation))
        if produit is None:
            absents.append(designation)
            continue
        if produit.quantity == quantite:
            inchanges += 1
            continue
        print(f"  {designation:38s} {produit.quantity:>5} → {quantite}")
        if ecrire:
            produit.quantity = quantite
            # Le produit repasse au-dessus de son seuil : l'alerte pourra se
            # redeclencher plus tard. Sans cette remise a zero, elle resterait
            # muette a la prochaine penurie.
            if quantite >= produit.seuil_alerte:
                produit.alert_sent = False
        changements += 1

    if ecrire and changements:
        db.commit()

    print()
    print(f"{changements} quantite(s) modifiee(s), {inchanges} deja a jour.")
    if absents:
        print(f"\nAbsents de la buvette — ignores ({len(absents)}) :")
        for nom_produit in absents:
            print(f"  · {nom_produit}")
        print("\n  La boutique HelloAsso fait foi sur ce qui existe. Pour en tenir")
        print("  compte, ajoutez-les d'abord a la boutique puis relancez la")
        print("  synchronisation depuis l'ecran Stock buvette.")
    return changements


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--fichier", required=True, type=Path, help="Classeur d'inventaire")
    parser.add_argument(
        "--feuille",
        default=None,
        help="Onglet a lire (par defaut : le dernier du classeur)",
    )
    parser.add_argument(
        "--appliquer",
        action="store_true",
        help="Ecrire reellement les quantites. Sans ce drapeau : simulation.",
    )
    args = parser.parse_args()

    if not args.fichier.is_file():
        raise SystemExit(f"Fichier introuvable : {args.fichier}")

    releve = lire_inventaire(args.fichier, args.feuille)
    if not releve:
        print("Aucun stock compte dans ce classeur.")
        return 0

    mode = "ECRITURE" if args.appliquer else "SIMULATION (rien n'est ecrit)"
    print(f"=== {mode} — {len(releve)} produit(s) compte(s) ===\n")

    db = SessionLocal()
    try:
        changements = appliquer(db, releve, ecrire=args.appliquer)
    finally:
        db.close()

    if not args.appliquer and changements:
        print("\nRelancer avec --appliquer pour ecrire.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
