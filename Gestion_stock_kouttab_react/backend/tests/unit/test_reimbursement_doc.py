"""Justificatif de remboursement : PDF et tableur remis a la comptabilite.

Ces fichiers sont la piece que la comptabilite archive ; ce sont eux qui
justifient le virement. Les tests verifient qu'on y retrouve tout ce que le
modele « NDF - Nom Prenom » impose : identite du benevole, une ligne par note,
les totaux, et le bloc « Apurement » qui dit comment le versement a ete fait.
"""

from __future__ import annotations

import zipfile
from datetime import date
from decimal import Decimal

import pytest

from app.services import reimbursement_doc


pytestmark = pytest.mark.unit


def _texte(pdf: bytes) -> str:
    """Texte du PDF, lu par `pypdf` (deja en dependance de test).

    reportlab encode ses flux en ASCII85 puis Flate : un extracteur maison
    coutait plus a maintenir qu'il ne rapportait, pour du code qui ne sert
    qu'ici.
    """
    import io

    from pypdf import PdfReader

    pages = PdfReader(io.BytesIO(pdf)).pages
    return "\n".join(page.extract_text() or "" for page in pages)


class _Note:
    def __init__(self, **kw):
        self.date_depense = kw.get("date_depense", date(2026, 8, 3))
        self.montant = Decimal(kw.get("montant", "42.50"))
        self.remboursement_deja_emis = Decimal(kw.get("avance", "0"))
        self.remise = Decimal(kw.get("remise", "0"))
        self.fournisseur = kw.get("fournisseur", "Metro")
        self.nature_charge = kw.get("nature_charge", "Alimentation")
        self.commentaires = kw.get("commentaires", "Gouter des enfants")
        self.evenement = kw.get("evenement")
        self.categorie = kw.get("categorie", "Courses")
        self.rattachement = kw.get("rattachement")
        self.pole = kw.get("pole", "Frais généraux")


class _Benevole:
    def __init__(self):
        self.nom = "Benfdila"
        self.prenom = "Omar"
        self.email = "omar@example.test"
        self.telephone = "06 01 02 03 04"
        self.username = "omar"


def _donnees(notes=None):
    return reimbursement_doc.DonneesJustificatif(
        benevole=_Benevole(),
        notes=notes if notes is not None else [_Note()],
        date_remboursement=date(2026, 8, 12),
        moyen="Virement bancaire",
        etablissement="Wise",
        approuve_par="DTC",
    )


# ---- PDF --------------------------------------------------------------------


def test_le_pdf_est_un_pdf_valide():
    octets = reimbursement_doc.construire_pdf(_donnees())
    assert octets.startswith(b"%PDF-")
    assert len(octets) > 800


def test_le_pdf_porte_l_identite_et_l_apurement():
    texte = _texte(reimbursement_doc.construire_pdf(_donnees()))
    for attendu in (
        "Benfdila",
        "Omar",
        "omar@example.test",
        "Apurement",
        "Virement bancaire",
        "Wise",
        "DTC",
    ):
        assert attendu in texte, f"{attendu!r} absent du PDF"


def test_le_pdf_contient_une_ligne_par_note():
    notes = [_Note(fournisseur="Metro"), _Note(fournisseur="Carrefour")]
    texte = _texte(reimbursement_doc.construire_pdf(_donnees(notes)))
    assert "Metro" in texte and "Carrefour" in texte


def test_le_rattachement_affiche_l_evenement_ou_la_categorie():
    """Meme arbitrage qu'au depot : evenement sous un pole EV, categorie sinon."""
    evenementielle = _Note(evenement="Gala de printemps", categorie=None)
    texte = _texte(reimbursement_doc.construire_pdf(_donnees([evenementielle])))
    assert "Gala de printemps" in texte

    texte = _texte(reimbursement_doc.construire_pdf(_donnees([_Note(categorie="Courses")])))
    assert "Courses" in texte


def test_le_total_du_pdf_deduit_avances_et_remises():
    notes = [_Note(montant="50.00", avance="20.00"), _Note(montant="10.00", remise="2.00")]
    texte = _texte(reimbursement_doc.construire_pdf(_donnees(notes)))
    # 30,00 + 8,00 = 38,00
    assert "38,00" in texte


# ---- Tableur ----------------------------------------------------------------


def test_le_xlsx_est_un_classeur_valide():
    octets = reimbursement_doc.construire_xlsx(_donnees())
    assert zipfile.is_zipfile(__import__("io").BytesIO(octets))


def test_le_xlsx_reprend_la_disposition_du_modele():
    """Libelles et emplacements du modele « NDF - Nom Prenom »."""
    import io

    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(reimbursement_doc.construire_xlsx(_donnees())))
    ws = wb.active

    assert ws["A1"].value == "Nom"
    assert ws["A2"].value == "Prénom"
    assert ws["A3"].value == "Mail"
    assert ws["A4"].value == "Téléphone"
    assert ws["B1"].value == "Benfdila"

    entetes = [ws.cell(row=10, column=c).value for c in range(1, 7)]
    assert entetes == [
        "Date",
        "Rattachement (évènement, activité..)",
        "Montant",
        "Fournisseur",
        "Nature charge",
        "Commentaires (Nombre de repas, …)",
    ]

    colonne_b = [c.value for c in ws["B"]]
    for libelle in (
        "Sous total :",
        "Remboursement déjà émis (espèce) :",
        "Remise (à préciser) :",
        "TOTAL :",
        "Remboursement émis le :",
        "Moyen :",
        "Etablissement :",
        "Approuvé par :",
    ):
        assert libelle in colonne_b, f"{libelle!r} absent du tableur"


def test_le_xlsx_ecrit_les_montants_en_nombres():
    """Des montants stockes en texte empecheraient toute somme dans Excel."""
    import io

    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(reimbursement_doc.construire_xlsx(_donnees())))
    ws = wb.active
    assert isinstance(ws["C11"].value, (int, float)), "le montant doit etre numerique"


def test_le_xlsx_liste_toutes_les_notes():
    import io

    from openpyxl import load_workbook

    notes = [_Note(fournisseur=f"Magasin {i}") for i in range(4)]
    wb = load_workbook(io.BytesIO(reimbursement_doc.construire_xlsx(_donnees(notes))))
    ws = wb.active
    fournisseurs = [ws.cell(row=11 + i, column=4).value for i in range(4)]
    assert fournisseurs == ["Magasin 0", "Magasin 1", "Magasin 2", "Magasin 3"]


def test_un_remboursement_sans_note_est_refuse():
    """Il n'y a rien a justifier : mieux vaut refuser que produire un document
    vide qui aurait l'air d'une piece valable."""
    with pytest.raises(Exception):
        reimbursement_doc.construire_pdf(_donnees([]))
