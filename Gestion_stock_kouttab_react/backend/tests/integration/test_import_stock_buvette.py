"""Alignement des quantités de la buvette sur un inventaire Excel.

Le script écrit dans le stock réel : ses deux risques sont d'écrire trop — une
cellule vide prise pour un zéro effacerait un stock existant — et d'écrire au
mauvais endroit, faute de rapprocher correctement des noms saisis à la main d'un
côté et par HelloAsso de l'autre.
"""

from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
import pytest
from sqlalchemy import select

RACINE = Path(__file__).resolve().parents[2]
if str(RACINE) not in sys.path:
    sys.path.insert(0, str(RACINE))

from scripts import importer_stock_buvette as importateur  # noqa: E402

from app.db.models import BuvetteProduct


def _classeur(tmp_path: Path, lignes: list[tuple[str, object]]) -> Path:
    """Reproduit la disposition réelle : en-têtes en ligne 5, données en 6."""
    classeur = openpyxl.Workbook()
    feuille = classeur.active
    feuille.title = "INV26-1"
    feuille.cell(row=5, column=2, value="Désignation")
    feuille.cell(row=5, column=8, value="11/7/2026")
    for index, (designation, stock) in enumerate(lignes, start=6):
        feuille.cell(row=index, column=1, value="Sucré")
        feuille.cell(row=index, column=2, value=designation)
        if stock is not None:
            feuille.cell(row=index, column=8, value=stock)
    chemin = tmp_path / "inventaire.xlsx"
    classeur.save(chemin)
    return chemin


@pytest.fixture()
def produits(db_session):
    articles = [
        BuvetteProduct(helloasso_tier_id=1, name="Palestine Cola", price_cents=100, quantity=3),
        BuvetteProduct(helloasso_tier_id=2, name="Oasis tropical", price_cents=100, quantity=10),
        BuvetteProduct(helloasso_tier_id=3, name="Miel d'acacia", price_cents=1500, quantity=7),
    ]
    db_session.add_all(articles)
    db_session.commit()
    return articles


def _quantite(db_session, nom: str) -> int:
    return db_session.execute(
        select(BuvetteProduct).where(BuvetteProduct.name == nom)
    ).scalar_one().quantity


def test_la_simulation_n_ecrit_rien(db_session, produits, tmp_path, capsys):
    """Le garde-fou : sans `--appliquer`, le stock ne bouge pas."""
    releve = importateur.lire_inventaire(_classeur(tmp_path, [("Palestine Cola", 50)]), None)

    importateur.appliquer(db_session, releve, ecrire=False)

    db_session.expire_all()
    assert _quantite(db_session, "Palestine Cola") == 3
    # ...mais le changement envisagé est annoncé, pour qu'on le relise.
    assert "50" in capsys.readouterr().out


def test_l_ecriture_applique_le_stock_compte(db_session, produits, tmp_path):
    releve = importateur.lire_inventaire(
        _classeur(tmp_path, [("Palestine Cola", 50), ("Oasis tropical", 67)]), None
    )

    importateur.appliquer(db_session, releve, ecrire=True)

    db_session.expire_all()
    assert _quantite(db_session, "Palestine Cola") == 50
    assert _quantite(db_session, "Oasis tropical") == 67


def test_une_cellule_vide_n_est_pas_un_zero(db_session, produits, tmp_path):
    """Le risque le plus coûteux.

    Une cellule vide signifie « pas compté », pas « stock nul ». L'écrire à zéro
    effacerait un stock réel, et personne ne s'en apercevrait avant la rupture.
    """
    releve = importateur.lire_inventaire(
        _classeur(tmp_path, [("Palestine Cola", None), ("Oasis tropical", 67)]), None
    )

    importateur.appliquer(db_session, releve, ecrire=True)

    db_session.expire_all()
    assert _quantite(db_session, "Palestine Cola") == 3, "le produit non compté ne bouge pas"
    assert _quantite(db_session, "Oasis tropical") == 67


def test_les_noms_se_rapprochent_malgre_accents_et_casse(db_session, produits, tmp_path):
    """Saisis à la main d'un côté, par HelloAsso de l'autre : ils divergent."""
    releve = importateur.lire_inventaire(
        _classeur(tmp_path, [("MIEL D'ACACIA", 12), ("oasis   Tropical", 5)]), None
    )

    importateur.appliquer(db_session, releve, ecrire=True)

    db_session.expire_all()
    assert _quantite(db_session, "Miel d'acacia") == 12
    assert _quantite(db_session, "Oasis tropical") == 5


def test_un_produit_absent_de_la_buvette_est_ignore(db_session, produits, tmp_path, capsys):
    """La boutique HelloAsso fait foi sur ce qui existe — mais on le signale."""
    releve = importateur.lire_inventaire(
        _classeur(tmp_path, [("Kinder Bueno (par 5)", 20), ("Palestine Cola", 50)]), None
    )

    importateur.appliquer(db_session, releve, ecrire=True)

    db_session.expire_all()
    assert _quantite(db_session, "Palestine Cola") == 50
    sortie = capsys.readouterr().out
    assert "Kinder Bueno" in sortie
    assert "ignor" in sortie.lower()


def test_le_stock_negatif_est_ramene_a_zero(db_session, produits, tmp_path):
    """Un inventaire ne rend jamais une quantité négative."""
    releve = importateur.lire_inventaire(_classeur(tmp_path, [("Palestine Cola", -4)]), None)

    importateur.appliquer(db_session, releve, ecrire=True)

    db_session.expire_all()
    assert _quantite(db_session, "Palestine Cola") == 0


def test_le_reapprovisionnement_rearme_l_alerte(db_session, produits, tmp_path):
    """Sans cette remise à zéro, l'alerte resterait muette à la prochaine pénurie."""
    article = db_session.execute(
        select(BuvetteProduct).where(BuvetteProduct.name == "Palestine Cola")
    ).scalar_one()
    article.seuil_alerte = 5
    article.alert_sent = True
    db_session.commit()

    releve = importateur.lire_inventaire(_classeur(tmp_path, [("Palestine Cola", 50)]), None)
    importateur.appliquer(db_session, releve, ecrire=True)

    db_session.expire_all()
    assert db_session.execute(
        select(BuvetteProduct).where(BuvetteProduct.name == "Palestine Cola")
    ).scalar_one().alert_sent is False
